"""
Output Formatter - Handles CSV, Excel, and Text output formats
"""
import csv
import json
from typing import List, Dict


class OutputFormatter:
    """Handles formatting STTM data for various output types"""
    
    @staticmethod
    def _format_value(value) -> str:
        """Format a value for output, handling complex types like lists and dicts"""
        if value is None:
            return ''
        if isinstance(value, (list, dict)):
            return json.dumps(value, default=str)
        return str(value)
    
    @staticmethod
    def _prepare_record(record: Dict) -> Dict:
        """Prepare a record for output by formatting complex values"""
        return {k: OutputFormatter._format_value(v) for k, v in record.items()}
    
    @staticmethod
    def to_csv(data: List[Dict], filepath: str):
        """Save data as CSV"""
        if not data:
            raise ValueError("No data to save")
        
        prepared_data = [OutputFormatter._prepare_record(r) for r in data]
        fieldnames = list(prepared_data[0].keys())
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(prepared_data)
    
    @staticmethod
    def to_excel(data: List[Dict], filepath: str):
        """Save data as Excel (requires openpyxl)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            raise ImportError(
                "openpyxl required for Excel output. "
                "Install with: pip install openpyxl"
            )
        
        if not data:
            raise ValueError("No data to save")
        
        prepared_data = [OutputFormatter._prepare_record(r) for r in data]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "STTM Mappings"
        
        headers = list(prepared_data[0].keys())
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_num, record in enumerate(prepared_data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                value = record.get(header, '')
                cell.value = value
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        for col_num, header in enumerate(headers, 1):
            max_length = len(header)
            for record in prepared_data:
                value = str(record.get(header, ''))
                max_length = max(max_length, len(value))
            
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[chr(64 + col_num) if col_num <= 26 else f"A{chr(64 + col_num - 26)}"].width = adjusted_width
        
        ws.freeze_panes = 'A2'
        
        wb.save(filepath)
    
    @staticmethod
    def to_text(data: List[Dict], filepath: str, format_type: str = 'markdown'):
        """Save data as formatted text"""
        if format_type == 'markdown':
            content = OutputFormatter._to_markdown(data)
        else:
            content = OutputFormatter._to_plain_text(data)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)
    
    @staticmethod
    def _to_markdown(data: List[Dict]) -> str:
        """Convert to markdown table"""
        if not data:
            return ""
        
        prepared_data = [OutputFormatter._prepare_record(r) for r in data]
        headers = list(prepared_data[0].keys())
        
        header_row = "| " + " | ".join(headers) + " |"
        separator_row = "| " + " | ".join(["---"] * len(headers)) + " |"
        
        data_rows = []
        for record in prepared_data:
            row_values = []
            for header in headers:
                value = str(record.get(header, ''))
                value = value.replace('|', '\\|')
                row_values.append(value)
            data_rows.append("| " + " | ".join(row_values) + " |")
        
        lines = [header_row, separator_row] + data_rows
        return "\n".join(lines)
    
    @staticmethod
    def _to_plain_text(data: List[Dict]) -> str:
        """Convert to plain text format"""
        if not data:
            return ""
        
        prepared_data = [OutputFormatter._prepare_record(r) for r in data]
        lines = []
        for i, record in enumerate(prepared_data, 1):
            lines.append(f"Record {i}:")
            lines.append("-" * 40)
            for key, value in record.items():
                lines.append(f"  {key}: {value}")
            lines.append("")
        
        return "\n".join(lines)
    
    @staticmethod
    def to_markdown_string(data: List[Dict]) -> str:
        """Convert mappings to markdown table format (returns string)"""
        return OutputFormatter._to_markdown(data)
